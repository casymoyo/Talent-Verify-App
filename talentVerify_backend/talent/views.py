from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import FileUploadParser
from .models import Department, Company, Role, Employee
import magic  
from openpyxl import load_workbook  
import pandas as pd  

class EmployeeUploadView(APIView):
    parser_classes = [FileUploadParser]  
    
    def post(self, request):
        if 'file' not in request.FILES:
            return Response({'error': 'No file uploaded'}, status=400)

        try:
            uploaded_file = request.FILES['file']

            mime_type = magic.Magic(mime=True).from_buffer(uploaded_file.read(1024), mime=True)
            uploaded_file.seek(0)  

            if mime_type not in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'text/csv']:
                return Response({'error': 'Unsupported file format. Please upload Excel (.xlsx) or CSV (.csv) files.'}, status=400)

            if mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                try:
                    wb = load_workbook(uploaded_file, data_only=True)
                    sheet = wb.active

                    data = sheet.values[1:]
                    df = pd.DataFrame(data, columns=[sheet.cell(row=1, col=c).value for c in range(1, sheet.max_column + 1)])
                except Exception as e:
                    return Response({'error': f"Error processing Excel file: {str(e)}"}, status=400)

                self.process_data(df)

            elif mime_type == 'text/csv':
                try:
                    df = pd.read_csv(uploaded_file)
                except Exception as e:
                    return Response({'error': f"Error processing CSV file: {str(e)}"}, status=400)

                self.process_data(df)

            return Response({'message': 'Employee data uploaded successfully'})

        except Exception as e:
            return Response({'error': str(e)}, status=400)
    
    def process_data(self, df):
        processed_data = []
        for index, row in df.iterrows():
            #company details
            company_name = row['Company Name']
            date_of_reg = row ['Date Of Registration']  
            reg_number = row['Registration Number']
            contact_person = row['Contact Person']
            number_of_employees = row['Number OF Employees']
            phonenumber = row['Phone Number']
            email = row['email']
            
            # departments details
            department_name = row['Department Name']
            
            # employee details
            employee_id = row['Employee ID']
            employee_name = row['Employee Name']
            
            # role details
            role_name = row['Role Name']
            date_started = row['Date Started']
            date_left = row['Date Left']
            duties = row['Duties']

            # Data validation
            if not company_name or not employee_id or not employee_name or not role_name:
                continue
            
            department, _ = Department.objects.get_or_create(name=department_name)

            company, _ = Company.objects.get_or_create(
                name=company_name,
                date_of_reg=date_of_reg, 
                reg_number=reg_number, 
                contact_person=contact_person,
                number_of_employees=number_of_employees,
                phonenumber=phonenumber,
                email=email,
                department=department
            )
                                
            role, _ = Role.objects.get_or_create(
                name=role_name,
                date_started=date_started,
                date_left=date_left,
                duties=duties
            )

            processed_data.append(Employee(
                employee_id=employee_id,
                company=company,
                name=employee_name,
                department=department,
                role=role
            ))

        with transaction.atomic():
            Employee.objects.bulk_create(processed_data)
            return redirect('success_url') 


